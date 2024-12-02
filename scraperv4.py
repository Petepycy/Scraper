import csv
import os
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

class CompanyRegistryExtractor:
    def __init__(self, url):
        self.url = url
        self.soup = self.get_soup()

    def get_soup(self):
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
        }
        response = requests.get(self.url, headers=headers)
        response.raise_for_status()
        return BeautifulSoup(response.content, "html.parser")

    def get_full_name(self, section):
        if section:
            full_name = section.find("h1")
            return full_name.text.strip() if full_name else None
        return None

    def extract_contact_email(self, section):
        try:
            email = section.find('div', class_='e-mail').find('span', class_='questo-paywall').text.strip()
            return email
        except AttributeError:
            return None

    def extract_contact_phone(self, section):
        try:
            phone = section.find('div', class_='phone').find('span').text.strip()
            return phone
        except AttributeError:
            return None

    def extract_contact_website(self, section):
        try:
            website = section.find('div', class_='site').find('span').text.strip()
            return website
        except AttributeError:
            return None

    def get_address(self, section):
        if section:
            address = section.find("span", class_="legal-address")
            return address.text.strip() if address else None
        return None

    def get_registry_details(self, section):
        details = {
            "NIP": None,
            "REGON": None,
            "KRS": None,
            "Forma prawna": None,
            "Data rejestracji na ALEO.com": None,
            "Adres rejestrowy": None,
            "Przedsiębiorca": None,
            "Małżeńska wspólność majątkowa": None,
            "Data rozpoczęcia działalności w CEIDG": None,
            "Status": None,
            "Main PKD Code": None,
            "PKD Codes": None
        }
        if section:
            rows = section.find_all("div", class_="registry-details__row")
            for row in rows:
                label = row.find("h3", class_="registry-details__row__label").text.strip()
                value = row.find("div", class_="registry-details__row__value").text.strip()
                if label in details:
                    details[label] = value

            pkd_codes = [code.text.strip() for code in section.find_all("p", class_="pkd-codes__code")]
            if pkd_codes:
                details["Main PKD Code"] = pkd_codes[0]
                details["PKD Codes"] = ", ".join(pkd_codes)
        return details

    def get_vat_status(self, section):
        try:
            vat_status = section.find("div", class_="vat-status").text.strip()
            return vat_status
        except AttributeError:
            return "VAT status not found."

    def extract_all(self):
        section = self.soup.find("app-registry-data-section")
        contact_section = self.soup.find("app-company-contact")
        registry_details = self.get_registry_details(section)
        return {
            "Nazwa pełna": self.get_full_name(section),
            "Adres siedziby": self.get_address(section),
            "NIP": registry_details.get("NIP"),
            "REGON": registry_details.get("REGON"),
            "KRS": registry_details.get("KRS"),
            "Forma prawna": registry_details.get("Forma prawna"),
            "Email": self.extract_contact_email(contact_section),
            "Telefon": self.extract_contact_phone(contact_section),
            "Strona WWW": self.extract_contact_website(contact_section),
            "Status": registry_details.get("Status"),
            "Data rejestracji na ALEO.com": registry_details.get("Data rejestracji na ALEO.com"),
            "Adres rejestrowy": registry_details.get("Adres rejestrowy"),
            "Przedsiębiorca": registry_details.get("Przedsiębiorca"),
            "Małżeńska wspólność majątkowa": registry_details.get("Małżeńska wspólność majątkowa"),
            "Data rozpoczęcia działalności w CEIDG": registry_details.get("Data rozpoczęcia działalności w CEIDG"),
            "Main PKD Code": registry_details.get("Main PKD Code"),
            "PKD Codes": registry_details.get("PKD Codes"),
            "VAT Status": self.get_vat_status(section),
            "Usunieta z KRS": self.get_removal_status(section)
        }

    def get_removal_status(self, section):
        removal_status = section.find("div", class_="removed-info")
        if removal_status:
            removal_text = removal_status.find("span", class_="removed-info__text")
            return removal_text.text.strip() if removal_text else None
        return None


class CompanyListExtractor:
    def __init__(self, url):
        self.url = url
        self.soup = self.get_soup()

    def get_soup(self):
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
        }
        response = requests.get(self.url, headers=headers)
        response.raise_for_status()
        return BeautifulSoup(response.content, "html.parser")

    def get_company_urls(self):
        company_urls = []
        section = self.soup.find("app-catalog-results")
        if section:
            companies = section.find_all("app-base-catalog-row")
            for company in companies:
                link = company.find("a", class_="catalog-row-first-line__company-name")
                if link:
                    company_urls.append("https://aleo.com/pl/" + link['href'])
        return company_urls


def save_to_csv(data, filename="extracted_data.csv"):
    if not data:
        print("No data to save.")
        return

    keys = data[0].keys()
    file_exists = os.path.isfile(filename)

    with open(filename, "a", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=keys)
        if not file_exists:
            writer.writeheader()
        writer.writerows(data)
    print(f"Data appended to {filename} successfully!")


def save_to_excel(data, filename="extracted_data.xlsx"):
    if not data:
        print("No data to save.")
        return

    keys = list(data[0].keys())
    file_exists = os.path.isfile(filename)

    if file_exists:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(keys)

    for row in data:
        sheet.append([row[key] for key in keys])

        # Adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Update or create a table with filters
    if "DataTable" in sheet.tables:
        table = sheet.tables["DataTable"]
        table.ref = f"A1:{get_column_letter(len(keys))}{sheet.max_row}"
    else:
        table = Table(displayName="DataTable", ref=f"A1:{get_column_letter(len(keys))}{sheet.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        sheet.add_table(table)

    workbook.save(filename)
    print(f"Data saved to {filename} successfully!")

if __name__ == "__main__":
    city = "gdynia"
    results_per_page = 25
    page_numbers = 3
    filename = "extracted_data.xlsx"

    for page_number in range(1, page_numbers + 1):
        if page_number == 1:
            main_url = f"https://aleo.com/pl/firmy?phrase=/{city}&count={results_per_page}"
        else:
            main_url = f"https://aleo.com/pl/firmy/{page_number}?phrase={city}&count={results_per_page}"
        list_extractor = CompanyListExtractor(main_url)

        company_urls = list_extractor.get_company_urls()
        all_data = []
        for url in company_urls:
            extractor = CompanyRegistryExtractor(url)
            data = extractor.extract_all()
            all_data.append(data)

        save_to_excel(all_data, filename=filename)
        save_to_csv(all_data, filename=filename.replace("xlsx", "csv"))
